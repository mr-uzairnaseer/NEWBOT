import sys
import os

try:
    import openpyxl
except ImportError:
    import subprocess
    print("openpyxl not found. Installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI Voicebot Requirements"

# Enable grid lines explicitly
ws.views.sheetView[0].showGridLines = True

# Styling definitions
font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="1F4E79")
font_normal = Font(name="Segoe UI", size=10, color="333333")

fill_title = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Corporate Blue
fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Steel Blue
fill_accent = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") # Very light grey-blue
fill_highlight = PatternFill(start_color="E6EEF8", end_color="E6EEF8", fill_type="solid") # Subtle blue highlight

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_title = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Set up Title Block
ws.merge_cells("A1:F2")
ws["A1"] = "ENTERPRISE AI VOICEBOT INTEGRATION - API REQUIREMENTS"
ws["A1"].font = font_title
ws["A1"].fill = fill_title
ws["A1"].alignment = align_title

# Set up headers
headers = [
    "Service Name", 
    "Role in Voicebot", 
    "Plan / Setup Cost", 
    "Estimated Usage Cost", 
    "Action Required from Employer", 
    "Sign-up / Access Link"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Data rows
data = [
    (
        "ElevenLabs",
        "Generative Text-to-Speech (TTS)\nProvides the highly realistic, natural-sounding human voice (emily) for outbound lead calls.",
        "Pro Plan ($99.00 / month)\n(Includes 500,000 characters/mo)",
        "Covered under the monthly plan credits. Overage is billed per character.",
        "None.\n(Already purchased by team. Developer has access to the account).",
        "https://elevenlabs.io"
    ),
    (
        "Deepgram",
        "Real-Time Speech-to-Text (STT)\nTranscribes the customer's voice instantly over the phone call, handling phone line noise & accents.",
        "Free Account Setup\n(Comes with free trial credits)",
        "~$0.0043 to $0.005 per minute of actual talk time (extremely cheap, pay-as-you-go).",
        "1. Create account.\n2. Add a credit card to activate the 'Pay-As-You-Go' tier.\n(This is a standard developer verification to unlock high concurrent call lines. No charges are made on card until free credits run out).",
        "https://console.deepgram.com"
    ),
    (
        "Groq",
        "Ultra-Fast LLM Hosting\nActs as the bot's 'brain' executing the Medicare script and deciding to Transfer/Drop.",
        "Free Account Setup\n(Generous free tier for testing)",
        "~$0.0005 per call (Virtually free: billed at $0.05 per 1 Million input tokens).",
        "1. Create account.\n2. Add a credit card to activate the 'On-Demand' tier.\n(Prevents the bot from hitting rate limits during outbound dialing. Card only billed for active usage, which is fractions of a cent).",
        "https://console.groq.com"
    )
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.font = font_normal
        cell.alignment = align_left
        cell.border = thin_border
        
        # Highlight first column
        if col_idx == 1:
            cell.font = font_bold
            cell.fill = fill_highlight
            cell.alignment = align_center

# Add summary / note block
ws.merge_cells("A8:F9")
ws["A8"] = "Executive Summary:\nThis API stack provides the absolute best real-time voicebot performance in the industry. By combining ElevenLabs' realistic voices, Deepgram's phone-optimized transcribing, and Groq's instantaneous Llama 3.1 brain, the bot will speak without awkward pauses, leading to higher qualification rates. Adding credit cards to Deepgram and Groq is a standard developer requirement to unlock concurrent outbound lines. Charges are strictly pay-as-you-go and negligible."
ws["A8"].font = Font(name="Segoe UI", size=9.5, italic=True, color="555555")
ws["A8"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws["A8"].fill = fill_accent

border_top_thick = Border(
    top=Side(style='thin', color='1F4E79'), 
    bottom=Side(style='thin', color='1F4E79'),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)
for col_num in range(1, 7):
    for r in [8, 9]:
        ws.cell(row=r, column=col_num).border = border_top_thick
        ws.cell(row=r, column=col_num).fill = fill_accent

# Set row heights
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 30
ws.row_dimensions[4].height = 65
ws.row_dimensions[5].height = 80
ws.row_dimensions[6].height = 65
ws.row_dimensions[8].height = 28
ws.row_dimensions[9].height = 28

# Adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        # Ignore merged title cell and summary cell for width calculation
        if cell.row in [1, 2, 8, 9]:
            continue
        val_str = str(cell.value or '')
        lines = val_str.split('\n')
        for line in lines:
            if len(line) > max_len:
                max_len = len(line)
    # Give columns some breathing room
    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Save workbook
output_path = "AI_Voicebot_Enterprise_Requirements.xlsx"
wb.save(output_path)
print(f"Excel file successfully generated at: {os.path.abspath(output_path)}")
